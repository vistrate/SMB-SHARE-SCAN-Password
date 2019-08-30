#!/usr/bin/python
# coding: utf8
########################################################################
__author__                       = "Vincent ROSET"
__copyright__ = "Appartient au propriétaire"
__credits__       = [""]
__license__       = "GPL"
__version__      = "0.1 07/06/2019"
########################################################################
import os, argparse, socket, re, csv, ipaddress, datetime, olefile, zipfile, sys
import docx, pptx
from pptx import Presentation
from openpyxl import Workbook, load_workbook

from sys import exit


from ipaddress import ip_address, ip_network
from contextlib import closing
from smb.SMBConnection import SMBConnection
from smb.smb_structs import ProtocolError

from smb.SMBConnection import SMBConnection
from smb.smb_structs import ProtocolError

from collections import deque
from logging import Logger
from win32 import win32security, win32api
from pathlib import Path

########LIBRARY########################################################

# Il est nécessaire d'installer avec pip ces deux bibliothèques
# la commande suivante se fait sur un terminal powershell:
# python.exe -m pip install pysmb --proxy 'http://user:password@fr-proxy03.knet.intra:3128' --user
# python.exe -m pip install pywin32 --proxy 'http://user:password@fr-proxy03.knet.intra:3128' --user
# python.exe -m pip install olefile --proxy 'http://user:password@fr-proxy03.knet.intra:3128' --user
# python.exe -m pip install python-docx --proxy 'http://user:password@fr-proxy03.knet.intra:3128' --user
# python.exe -m pip install python-pptx --proxy 'http://user:password@fr-proxy03.knet.intra:3128' --user
# python.exe -m pip install openpyxl --proxy 'http://user:password@fr-proxy03.knet.intra:3128' --user
# python.exe -m pip install oletools --proxy 'http://user:password@fr-proxy03.knet.intra:3128' --user
#####################GLOBAL Variables###################################################

client_machine_name = socket.gethostname()
domain_name = 'KNET.INTRA'
use_ntlm_v2 = True
PasswordBase = {}
AccessList = {}
verboseOption = False
scanContentOption = False
regex_txt = '(.+)\.(txt|init|conf)$'
regex_office = '(.+)\.(xls|xlsx|xlsm|xml|doc|docm|docx|ppt|pptx|pptm)$'
regex_login = r'(.*)knet\\(.*)$'
regex_password = r'(.*)password(.*)$'
regexFileNameList = []
regexContentList = []
depth = 5
output_filename = 'stdout'
searchlastwriterOption = False
portSMB = 139


########################################################################

#Génère un range d'adresse IP sur le dernier octet en IPv4
#Input : Une adresse IPv4 au format string, avec l'un des trois 
# derniers octets contenant un -, et le numéro de l'octet concerné
# Exemple : 10.0.0.1-3 -> 10.0.0.1, 10.0.0.2, 10.0.0.3
#Output : liste d'adresses IPv4 correspondant à une plage
def getIPListFromRange(addr, bytePosition): 
	ip_parts = []
	min,max = (0, 0)
	for part in addr.split("."):
		if "-" in part:
			min, max = part.split("-")
			ip_parts.append((min,max))
		else:
			ip_parts.append(part)

	min = int(min)
	max = int(max)
	if max <= min : return []

	ip_list = []
	c = ""
	if bytePosition == 4 :
		c = ip_parts[0] + '.' + ip_parts[1] + '.' +  ip_parts[2] + '.'
		for i in range(min, max +1) :
			s = str(i)
			ip_list.append(c + s)
	elif bytePosition == 3:
		c = ip_parts[0] + '.' + ip_parts[1] + '.'
		for i in range(min, max + 1) :
			for k in range(0, 256):
				s0 = str(i)
				s1 = str(k)
				ip_list.append(c + s0 + '.' + s1)
	elif bytePosition == 2:
		c = ip_parts[0] + '.'
		for i in range(min, max + 1) :
			for k in range(0, 256) :
				for j in range(0,256) :
					s0 = str(i)
					s1 = str(k)
					s2 = str(j)
					ip_list.append(c + s0 + '.' + s1 + '.' + s2)
	return ip_list




# A partir d'un SID d'un fichier, retourne le nom de l'utilisateur présent
# sur le domaine du serveur où stocké le fichier
# Input = 
	# - sid d'un fichier (en string)
	# - adresse IP du serveur distant (où est localisé le fichier)
# Output = retourne le nom du propriétaire du fichier au format Domain\nom
def get_owner(sid, remote_server):
	user_sid  = win32security.ConvertStringSidToSid(sid)
	principle, domain, type = win32security.LookupAccountSid(remote_server, user_sid)
	user_fq = domain + "\\" + principle
	return user_fq


# Retourne sur des fichiers office (word, excel et powerpoint)
# le nom de la dernière personne a voir effectuée dessus
# des enregistrements.

# Input =
	# - nom du chemin du fichier
	# - extension du fichier
# Output = 
	# le nom de l'auteur du dernier enregistrement
	# et si rien n'est trouvé une chaine vide
def findLastAuthor(filename, suffix):
	if suffix in ['.doc', '.xls', '.ppt']:
		return get_last_author_doc_xls_ppt(filename)
	elif suffix in ['.pptx', 'pptm']:
		return get_last_author_pptx_pptm(filename)
	elif suffix in ['.xlsx', 'xlsm']:
		return get_last_author_xlsx_xlsm(filename)
	elif suffix == '.docx':
		return get_last_author_docx(filename)
	else:
		return ""

# Les autres fonctions ci-dessous sont associés à la dernière 
# fonction. Il y a des fonctions différentes par extension
# car les bibliothèques associés y sont différentes

def get_last_author_doc_xls_ppt(filename):
	global ole
	last = ""
	try:
		ole = olefile.OleFileIO(filename)
		meta = ole.get_metadata()
		last = meta.last_saved_by
		last = last.decode('utf-8')
	except:
		last = ""
	finally:
		ole.close()
		return last


def get_last_author_docx(filename):
	path = Path(filename)
	last = ""
	if zipfile.is_zipfile(path):
		try:
			docxFile = docx.Document(path)
			docxInfo = docxFile.core_properties
			last = getattr(docxInfo, 'last_modified_by')
		except :
			last = ""
		finally:
			return last
	else:
		return ""


def get_last_author_pptx_pptm(filename):
	path = Path(filename)
	last = ""	
	if zipfile.is_zipfile(path):
		try:
			prs = Presentation(path)
			prsInfo = prs.core_properties
			last = getattr(prsInfo, 'last_modified_by')
		except :
			last = ""
		finally:
			return last
	else:
		return ""

def get_last_author_xlsx_xlsm(filename):
	path = Path(filename)
	last = ""
	if zipfile.is_zipfile(path) :
		try:
			wb = Workbook()
			wb = load_workbook(path)
			last = wb.properties.lastModifiedBy
		except :
			last = ""
		finally:
			return last
	else:
		return ""

# Vérifie la validité d'un hostname
#Input : hostname au format string
#Output : True si le hostname est valide, false le cas échéant
def is_valid_hostname(hostname):
	if len(hostname) > 255:
		return False
	if hostname[-1] == ".":
		hostname = hostname[:-1] # strip exactly one dot from the right, if present
	allowed = re.compile("(?!-)[A-Z\d-]{1,63}(?<!-)$", re.IGNORECASE)
	return all(allowed.match(x) for x in hostname.split("."))


# Groupe de fonction utilisé pour chercher lors de la création d'une plage
# l'octet concerné par un range
# Input : adresse IP
# Output: True si le deuxième ou troisième ou dernier octet est matché
def isIPRangeLastByte(addr):
	m = re.match(r"^(\d{1,3})\.(\d{1,3})\.(\d{1,3})\.(\d{1,3})-(\d{1,3})$", addr)
	return bool(m) and all(map(lambda n: 0 <= int(n) <= 255, m.groups()))

def isIPRangeThirdByte(addr):
	m = re.match(r"^(\d{1,3})\.(\d{1,3})\.(\d{1,3})-(\d{1,3}).(\d{1,3})$", addr)
	return bool(m) and all(map(lambda n: 0 <= int(n) <= 255, m.groups()))	
	
def isIPRangeSecondByte(addr):
	m = re.match(r"^(\d{1,3})\.(\d{1,3})-(\d{1,3}).(\d{1,3})\.(\d{1,3})$", addr)
	return bool(m) and all(map(lambda n: 0 <= int(n) <= 255, m.groups()))
		


#Selon la doc s'il n'y a pas besoin d'authentification, on peut mettre ce que l'on veut dans
# user, password, client_machine_name, server_name et le domaine
#Input = IP/Hostname d'un hôte et le numéro de port du service SMB (139 par défaut)
#Output = True si le service nécessite une authentification False le cas échéant
def isRequiredAuthentication(host, port=139):
	c = None
	try:
		c = connectToSMBService("guekljst", "bbbb", "sdf", "sdflva",host, "mmmd", port, True) 
		c.close()
	except:
		pass
	finally:
		
		return (c == None)


# Vérifie si un port TCP est ouvert sur un hôte
# Input = IP/Hostname d'un hôte et le numéro de port
# Output = True si le port est ouvert, False dans le cas échéant
def isPortOpen(ip, port):
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(1)
        try:
                s.connect((ip, int(port)))
                s.shutdown(socket.SHUT_RDWR)
                return True
        except:
                return False
        finally:
                s.close()

# Se connecte à un service SMB.
# Input : prend en entrée le user, password, le nom de la machine cliente (une valeur quelconque peut être mise),
# le hostname du serveur, l'adresse IP du serveur, le nom de domaine, le port SMB utilisé (139, 445),
# le choix entre ntlm v1 et v2 (True indiquera une utilisation de ntlm v2 et False une utilisation de ntlm v1)
# Output : retourne un objet SMB de connexion
def connectToSMBService(user, password, client_machine_name, server_name, server_ip, domain_name, portSMB=139, use_ntlm_v2=True):
	
	if not isPortOpen(server_ip, portSMB): 
		print("Service SMB fermé sur %s %d" % (server_name, portSMB))
		return None
		
	try:
		conn = SMBConnection(user, password, client_machine_name, server_name, domain_name, use_ntlm_v2)
		connect_result = conn.connect(server_ip, portSMB)
		if connect_result is False:
			raise ProtocolError("Echec de connexion au service SMB")
	except ProtocolError as e:
		logger.debug("Exception lors de la connexion à %s: %s" % (server_name, e))
		raise exceptions.ConnectException("Connexion à %s avortée, mauvais mot de passe ?" % server_name)
		return None
	else:
		return conn

# Ferme une connection SMB active
# Input : Objet de connexion SMB
# Output : Rien
def closeSMBConnection(connection):
	if connection == None:
		sys.exit("L'objet connection est vide")
	else:
		connection.close()



# Convertit une valeur floattante correspondant au nombre de
# secondes écoulés depuis 1970-01-01 00:00:00
# Input : nombre de secondes en flottant
# Output : date au format string 1970-01-01 00:00:00
def getTime(time):
	timestamp = datetime.datetime.fromtimestamp(time)
	d = timestamp.strftime('%Y-%m-%d %H:%M:%S')
	return d

def removeDoubleBackslash(str):
	s = str.replace('\\\\', '\\')
	return s

def isRegexMatched(pattern, filename):
	prog = re.compile(pattern, re.IGNORECASE)
	return prog.search(filename)



#Recherche sur un partage réseau les fichiers dont le nom de fichier matche
# avec une regex
# A partir d'identifiants de connexion et d'une adresse de partage réseau
# et d'un filtre regex.

#Input : 
	# - user 
	# - password
	# - nom de machine cliente 
	# - hostname du serveur SMB
	# - adresse IP du serveur SMB
	# - Liste de regex
	# - option verbose
	# - portSMB (par défaut 139)
	# - nom de domaine active directory
	# - utilisation ou non de SMB v2 (sinon SMB v1)
	# - recherche ou non de l'auteur du dernier enregistrement (prends du temps)
	# - profondeur max du scan (si -1 est positionné il n'est pas pris en compte)

# Output:
	# Liste de nom de fichiers contenant
	# - adresse IP de l'hôte
	# - hostname de l'hôte
	# - date de création
	# - propriétaire du fichier
	# - nom du fichier
	# - emplacement du fichier
	# - taille du fichier
	# - dernier utilisateur à l'avoir modifié
	# - regex matché


def SearchFileName(user, password, client_machine_name, server_name, server_ip, regexList, verbose,
	portSMB, domain_name, use_ntlm_v2, giveLastAuthor, depth_max):
	
	# ouverture de la connexion
	conn = connectToSMBService(user, password, client_machine_name, server_name, server_ip, domain_name,
	portSMB, use_ntlm_v2)
	
	if conn == None:
		return []
	
	
	list_of_File =  []
	
	suffix_list = ['.xls', '.xlsm', '.xml','.xlsx','.doc', '.docx', '.ppt', '.pptx', '.pptm']
	
	#on traite tous les dossiers partagés accessibles
	# partage réseau == sharedDevice
	shares = conn.listShares(timeout=30)
	for share in shares:
		if share.isSpecial or share.name in ['NETLOGON', 'SYSVOL']:
			continue
		else:
			if verbose :
				print("**************%s*************"%(share.name))
			queue = deque()
			path = '\\'

			queue.append((share, path))
			while queue :
				t = queue.pop()
				path = t[1]
				p = Path(path)
				if depth_max != -1:
					if len(p.parts) > depth_max :
						continue
				try:
					listOfFilesOnDirectory = conn.listPath(t[0].name, path)
				except :
					#il y a des exceptions sur les droits d'accès nous ne nous en occuperons pas
					#pour le moment
					continue
				else:	
					for sharedFile in listOfFilesOnDirectory:
						if sharedFile.filename != '..' and sharedFile.filename != '.':
							temp_path = path + sharedFile.filename
							if sharedFile.isDirectory:
								temp_path +=  '\\'
								queue.append((share, temp_path))
							else :
								# zone dans lequel un fichier régulier va être traité

								p = None
								suffix = ""
								
								for regex in regexList:
									if isRegexMatched( regex, sharedFile.filename) != None :
										listOfAttributesFile = conn.getSecurity(t[0].name, path)
										
										dico = dict()
										dico['ip'] = server_ip
										dico['hostname'] = server_name
										dico['create_time'] = getTime(sharedFile.create_time)
										#le sid doit être au format string pour qu'il soit pris en compte
										dico['owner'] =  get_owner(str(listOfAttributesFile.owner), server_ip)
										dico['name'] = sharedFile.filename
										path_temp =  str(server_ip) + '\\' + share.name + temp_path
										dico['path'] = '\\\\' + removeDoubleBackslash(path_temp)
										dico['length'] = sharedFile.file_size
										p = Path(dico['path'])
										if(giveLastAuthor):
											suffix = p.suffix
											if suffix in suffix_list:
												dico['last_user'] = findLastAuthor(dico['path'] , suffix)
												
										dico['regex_fileName'] = regex
										list_of_File.append(dico)
										if verbose :
											print(dico)
										
			queue.clear()

	# Fermeture de la connexion
	closeSMBConnection(conn)
	return list_of_File


#Input : 
	# - user 
	# - password
	# - nom de machine cliente 
	# - hostname du serveur SMB
	# - adresse IP du serveur SMB
	# - Liste de regex sur le nom de fichier
	# - Liste de regex sur le contenu du fichier
	# - option verbose
	# - portSMB (par défaut 139)
	# - nom de domaine active directory
	# - utilisation ou non de SMB v2 (sinon SMB v1)
	# - recherche ou non de l'auteur du dernier enregistrement (prends du temps)
	# - profondeur max du scan (si -1 est positionné il n'est pas pris en compte)
	
# Output:
	# Liste de nom de fichiers contenant
	# - adresse IP de l'hôte
	# - hostname de l'hôte
	# - date de création
	# - propriétaire du fichier
	# - nom du fichier
	# - emplacement du fichier
	# - taille du fichier
	# - dernier utilisateur à l'avoir modifié 
	# - regex matché
	
	# Ligne dans un fichier dont le contenu correspond 
	# à un motif recherché

def SearchFileContent(user, password, client_machine_name, server_name, server_ip, regexListNameFile,
	verbose, regexListContentFile,portSMB, domain_name, use_ntlm_v2,
	giveLastAuthor, depth_max):
	
	# ouverture de la connexion
	conn = connectToSMBService(user, password, client_machine_name, server_name, server_ip, domain_name,
	portSMB, use_ntlm_v2)
	
	if conn == None:
		return []
	
	list_of_File =  []
	
	suffix_list = ['.xls', '.xlsm', '.xml','.xlsx','.doc', '.docx', '.ppt', '.pptx', '.pptm']
	
	#on traite tous les dossiers partagés accessibles
	# partage réseau == sharedDevice
	shares = conn.listShares(timeout=30)
	for share in shares:
		if share.isSpecial or share.name in ['NETLOGON', 'SYSVOL']:
			continue
		else:
			if verbose :
				print("**************%s*************"%(share.name))
			queue = deque()
			path = '\\'

			queue.append((share, path))
			while queue :
				t = queue.pop()
				path = t[1]
				p = Path(path)
				if depth_max != -1:
					if len(p.parts) > depth_max :
						continue
				try:
					listOfFilesOnDirectory = conn.listPath(t[0].name, path)
				except :
					#il y a des exceptions sur les droits d'accès nous ne nous en occuperons pas
					#pour le moment
					continue
				else:	
					for sharedFile in listOfFilesOnDirectory:
						if sharedFile.filename != '..' and sharedFile.filename != '.':
							temp_path = path + sharedFile.filename
							if sharedFile.isDirectory:
								temp_path +=  '\\'
								queue.append((share, temp_path))
							else :
								# zone dans lequel un fichier régulier va être traité

								p = None
								suffix = ""
								
								for regex in regexListNameFile:
									if isRegexMatched( regex, sharedFile.filename) != None :
										listOfAttributesFile = conn.getSecurity(t[0].name, path)
										
										dico = dict()
										dico['ip'] = server_ip
										dico['hostname'] = server_name
										dico['create_time'] = getTime(sharedFile.create_time)
										#le sid doit être au format string pour qu'il soit pris en compte
										dico['owner'] =  get_owner(str(listOfAttributesFile.owner), server_ip)
										dico['name'] = sharedFile.filename
										path_temp =  str(server_ip) + '\\' + share.name + temp_path
										dico['path'] = '\\\\' + removeDoubleBackslash(path_temp)
										dico['length'] = sharedFile.file_size
										p = Path(dico['path'])
										if(giveLastAuthor):
											suffix = p.suffix
											if suffix in suffix_list:
												dico['last_user'] = findLastAuthor(dico['path'] , suffix)
												
										dico['regex_fileName'] = regex
										
										for regex_bis in regexListContentFile:
											MatchesContent = {}
											try:
												f = open(p, 'r', errors='ignore')
											except:
												continue
												
											i = 1
											for x in f:
												print(i)
												print(x)
												i = i + 1
												res = re.findall(regex_bis, x, flags=re.IGNORECASE)
												if  res != []:
													greped = 'Ligne : ' + str(i) + ' '
													for it in res:
														greped += str(it) + ' '
													MatchesContent['regex_ContentFile'] = regex_bis
													MatchesContent['MatchesElement'] = greped
												i = i + 1
											f.close()
											if MatchesContent != {}:
												dico['MatchesContent'] = MatchesContent
												
										list_of_File.append(dico)
										if verbose :
											print(dico)
								
											
										
										
			queue.clear()

	# Fermeture de la connexion
	closeSMBConnection(conn)
	return list_of_File

# Vérifie si le login a une forme correcte (pas de caractère spéciaux)
# et si le mot de passe n'est pas vide
# Input :
		# - login (chaîne de caractère)
		# - password (chaîne de caractère)
# Output :
		# True si le couple login/mot de passe est valide
		# False si le couple login/mot de passe est invalide
def checkAccount(user, password):
	userprog = re.compile('^[0-9a-zA-Z\-]+$')
	passwordprog = re.compile('.+')
	return userprog.match(user) is not None and passwordprog.match(password) is not None



# Retourne une liste de compte SMB provenant d'une liste de login/mot de passe en ligne de commande
# ou provenant d'un fichier au format csv
# le user ne doit pas contenir de caractères spéciaux
# le mot de passe ne doit pas être vide
# Les logins/mots de passes ne respectant pas la nomenclature seront ignorés
# L'entête aussi est ignorée (en option)
# Input :
	# - liste de login provenant de la ligne de commande
	# - liste de password provenant de la ligne de commande
	# - chemin du fichier contenant la liste des logins mots de passes
# Output:
	# Liste des logins/Mot de passe au format [ (login, mot de passe) , ...]
def getSMBaccountFromInputForUser(Path_of_fileAccount = "", List_of_users_from_prompt=[], List_of_passwords_from_prompt=[], has_header = True):

	list_of_account = []

	#Traitement du fichier csv
	if Path_of_fileAccount != "" :
		with open(Path_of_fileAccount, newline='') as f:
			reader = csv.reader(f, delimiter=';')
			try:
				# s'il y a un header (IP/Login/Mot de passe) il n'est pas pris en compte
				if(has_header):
					next(f)
				for row in reader:
					if checkAccount(row[1], row[2]):
						list_of_account.append((row[1], row[2]))
					else:
						print("Login/Mot de passe : %s %s invalide" %(row[1], row[2]))
						
			except csv.Error as e:
				print("file {%s}, line {%i}: {%s}" % (Path_of_fileAccount, reader.line_num, e))
			except IndexError:
				print("Nombre de login différent du nombre de mot de passe !")
				pass
				
	#Traitement des paramètres en ligne de commande	
	if List_of_users_from_prompt != [] and List_of_passwords_from_prompt != [] :

		if len(List_of_users_from_prompt) != len(List_of_passwords_from_prompt) :
			print("Nombre de login différent du nombre de mot de passe !")
		else:
			for i in range(0, len(List_of_users_from_prompt)):
				if checkAccount(List_of_users_from_prompt[i], List_of_passwords_from_prompt[i]) :
					list_of_account.append((List_of_users_from_prompt[i], List_of_passwords_from_prompt[i]))
					
	
	return list_of_account

# Prends en entrée une IP/Hostnames/Networks/range
# et transmets les bonnes IPs associés
# Input = Une IP/Hostnames/Networks/range au format string
# Output = liste d'IPs au format string
def checkIPList(addr):
	list_of_ip = []

	try:
		if "/" in addr : 
			network = ipaddress.ip_network(addr)
			for it in network.hosts():
				list_of_ip.append(it)
		elif isIPRangeLastByte(addr) or isIPRangeThirdByte(addr) or isIPRangeSecondByte(addr):
			position = 0
			if isIPRangeSecondByte(addr) : 
				position = 2
			elif isIPRangeThirdByte(addr):
				position = 3
			elif isIPRangeLastByte(addr):
				position = 4
			
			t = getIPListFromRange(addr, position)
			if t != [] :
				for it in t:
					network = ipaddress.ip_address(it)
					list_of_ip.append(network)
		else:
			addr_bis = addr
			if is_valid_hostname(addr):
				addr_bis = socket.gethostbyname(addr)
			network = ipaddress.ip_address(addr_bis)
	except (OSError, ValueError) as e:
		print(e)
		pass
	else:
		list_of_ip.append(network)
		
	list_of_addr = []
	if list_of_ip != []:
		for it in list_of_ip:
			list_of_addr.append(str(it))
	
	return list_of_addr


# Prends une ip/Hostnames/Networks en ligne de commande
# ou le chemin d'un fichier au format csv contenant une telle liste de ces éléments
# Une vérification est effectuée pour s'assurer de la bonne nomenclature
# des IP/Hostnames/Networks entrées.
# Retourne la liste correspondant au bon username dans le cas d'un fichier
# Input = string contenant un ip/Hostnames/Networks/range
# ou (inclusif)  un chemin de fichier contenant une telle liste 
# on indique si le fichier csv a un header ou non
# Output = Liste contenant toutes les adresses IP associées au différents 
# IP/Hostnames/Networks
def getListOfIPsFromInputForUser(path_of_IPList= "", IP_from_prompt="", login = "", has_header=True):
	
	list_of_ip = []

	#Cas d'un fichier contenant les adresses, login etc
	if path_of_IPList != "" and login != "" :
		filename = path_of_IPList
		with open(filename, newline='') as f:
			reader = csv.reader(f, delimiter=';')
			if (has_header):
				next(f)
			try:
				# on ne prends que la première colonne
				for row in reader:
					if (row[1] == login) :
						list_of_ip += checkIPList(row[0])
			except csv.Error as e:
				print('file {}, line {}: {}'.format(filename, reader.line_num, e))

	if IP_from_prompt != "":
		list_of_ip += checkIPList(IP_from_prompt)
	
	return list_of_ip
	

# Ajoute ou modifie une entrée existante dans un dictionnaire d'accès
# Ce dictionnaire indique quel utilisateur a accès à telle liste d'IPs.
# Chaque sous-liste correspond à un utilsateur.
# Le nombre d'utilisateur, doit être le même que le nombre de sous-liste.
# Car l'on va traiter des lignes de la manière suivante
# admin, adresse IP1, adresse IP2 
# admin, adresse IP...
# toto, adresse IP...
# Il n'est pas possible de rajouter un doublon ex : redonner à admin la même adresse IP
#Exemple :
# L'utilisateur admin a pour liste ['10.0.0.3', '10.0.0.4']
# L'utilisateur toto a pour liste ['10.0.0.1', '10.0.0.2', '10.0.0.54']
# User = 'admin'
# HostList = [['10.0.0.3', '10.0.0.4'], ['10.0.0.1', '10.0.0.2', '10.0.0.54']]
# Le dictionnaire a pour valeur après ajout: {'admin': {'10.0.0.3', '10.0.0.4'}, 'toto': {'10.0.0.1', '10.0.0.2', '10.0.0.54'}}

#Input : 
# un login (string login)
# une liste de sous-liste d'adresses IP
# un dictionnaire existant
# Output :
#  modifie le dictionnaire entrée en variable globale
def addEntrieAccessList(username, HostList):
	global AccessList
	i = 0
	for i in range(0, len(HostList)):
		# on teste si l'utilisateur est déjà présent dans le dictionnaire
		if username in AccessList:
			old_list = AccessList[username]
			other_list = HostList[i]
		# Dans le cas où l'entrée existe déjà dans le dictionnaire,
		# on vérifie si l'adresse IP à ajouter n'est pas déjà présente
			for ip in other_list:
				if ip not in old_list:
					old_list.append(ip)
			AccessList[username] = old_list
		else:
			AccessList[username] = HostList[i]
		i = i + 1
				
# Créé un dictionnaire avec pour clé un login et valeur un mot de passe
# Exemple : admin a pour mot de passe lsnfpjepof
		# toto a pour mot de passe ejfpefkeofd
	# donc 
		# UserList = ['admin', 'toto']
		# PasswordList = ['lsnfpjepof', 'ejfpefkeofd']
		# Le dictionnaire aura pour valeur : {'admin': 'lsnfpjepof', 'toto': 'ejfpefkeofd'}
	# Le nombre d'utilisateur, doit être le même que le nombre de mot de passe car 
# l'on va traiter les éléments lignes à ligne exemple:
	# admin ; mdp1
	# toto ; mdp2
	# user x; mdp y 
# A noter que si une ligne avec un doublon sur un user est trouvé, la nouvelle ligne écrasera 
# la valeur de l'ancien mot de passe

# Input : 
		# - une liste d'utilisateur en chaîne de caractère
		# - une liste de mot de passes en chaîne de caractère
# Output :
		# un dictionnaire associant login et mot de passe
def createPasswordBase(UserList, PasswordList):
	global PasswordBase
	if len(UserList) != len(PasswordList):
		print("erreur, le nombre d'utilisateurs n'est pas égale"  +
			" au même nombre de mots de passe")	
		return dict()
	i = 0
	for user in UserList:
		PasswordBase[user] = PasswordList[i]
		i += 1



# Retourne le mot de passe correspondant à l'utilisateur
# dans le dictionnaire des mots de passe.
# Input :
	# - Dictionnaire des mots de passes (structure dict)
	# - Un utilisateur en chaine de caractère
# Output :
	# Le mot de passe associé à un utilisateur en string.
	# s'il existe
def givePasswordFromUser(user):
	global PasswordBase
	if PasswordBase.get(user) != None:
		return PasswordBase[user]
	else :
		print("L'utilisateur %s n'existe pas dans la base " % (user) +
		"de mot de passe.")
		return ""


#Retourne un range d'adresse IP associé à un utilisateur.
# LE range est recherché dans le dictionnaire d'accès
# (variable globale AccessList)
# Input : 
# - chaîne de caractère user
# Output:
# - Liste d'adresse IP ou une liste vide
def getListOfIPFromUserAccess(user):
	global AccessList
	if AccessList.get(user) != None:
		return AccessList[user]
	else :
		print("L'utilisateur %s n'existe pas dans la liste " % (user) +
		"des accès.")
		return ""
	

# Traite les arguments entrées sur le prompt
# input : Argument du prompt
# output : modifie les dictionnaires de bases de mot de passe et lié aux access list
def handleArgument():

	parser = argparse.ArgumentParser()

	#Traitement des arguments en ligne de commande concernant les comptes
	# il n'est possible de rentrer qu'un seule compte en ligne de commande
	parser.add_argument('-u', '--user', help='Login d\'un utilisateur', type=str, nargs='?')
	parser.add_argument('-p', '--password', help='Mot de passe d\'un utilisateur', type=str, nargs='?')

	#Traitement des arguments en ligne de commande concernant les IP/Hostnames/Networks
	parser.add_argument('-ip', help='Adresse IP, Plage d\'IP, hostname et sous-réseau. Il est possible \
	d\'utiliser des ranges sur les 3 derniers octets. Exemple : 10.0.0.1-2 -> \
	10.0.0.1 , 10.0.0.2 ou une liste de plusieurs arguments à la suite Exemple: 10.0.0.1 localhost 10.0.0.0/24',
	type=str, nargs='*')

	#traitement en argument du chemin du fichier d'hôtes et login/mot de passe
	parser.add_argument('-f', help='Chemin d\'un fichier contenant des adresses IP/Hostnames/Networks \
	et le compte SMB associé (login / mot de passes). Si des entrées sont mis en ligne\
	de commande. Par contre s\'il n\'y a pas d\'entrées en ligne de commande, le fichier peut suffire.\
	Des adresses IP, des plages d\'ip, des subnets et des hostnames peuvent être mis dedans dans la première colonne.\
	Puis la seconde colonne contient le login, et la troisième colonne le mot de passe associé.\
	Le fichier doit être au format csv avec un ; comme séparateur. \
	Exemple : -f \'toto.csv\' \
	Format d\'un fichier en entrée contenant les hôtes et login/mot de passe : \
	*******************************************************\
	* Liste des IP/Hostnames/Networks * Login/Mot de passe*\
	*******************************************************\
	10.10.0.1                         * toto / admin\
	www.klepierre.commande            * john / smith\
	10.0.0.0/24                       * toto / admin\
	192.168.0.1-55                    * titi / 12345\
	..............',type=str, nargs='?')
	
	parser.add_argument('-v', '--verbose', action='store_true', default=False,
	help='Permet de suivre l\'avancement du scan')

	parser.add_argument('-c', '--scancontent', action='store_true', default=False,
	help='Scanne aussi le contenu du fichier à la recherche de motif')

	parser.add_argument('-regexfilename', help='Prends une liste de regex sur les noms de fichiers',type=str, nargs='*')
	
	parser.add_argument('-regexoffice', action='store_true', default=False, 
	help='Utiliser la regex sur les fichiers d\'extension office : \
	xls, xlsm, xml, xlsx, doc, docx, ppt, pptx, pptm')
	parser.add_argument('-regextxt', action='store_true', default=False, 
	help='Utiliser la regex sur les fichiers d\'extension texte : \
	txt, init, conf')
	
	parser.add_argument('-regexcontent', help='Prends une liste de regex sur les contenus des fichiers',type=str, nargs='*')
	
	parser.add_argument('-maxdepth', help='Définit la profondeur maximale du scan',type=int, nargs='?')
	
	parser.add_argument('-output', '-o', help='Chemin du fichier de sortie',type=str, nargs='?')

	parser.add_argument('-searchlastwriter', help='Indique si le scan recherche aussi le dernier utilisateur \
	a avoir écrit dans le fichier',action='store_true', default=False)

	parser.add_argument('-ntlmv1', help='Indique si ntlm version 1 doit être utilisé', action='store_true', default=False)
	
	parser.add_argument('-portsmb', help='Indique le numéro de port SMB (139 ou 445)',type=int, nargs='?', default=139, choices=[139, 445])
	
	parser.add_argument('-domainname', help='Indique le nom de domaine a utiliser.',type=str, nargs='?', default='knet.intra')
	
	parser.add_argument('-clientmachinename', help='Indique le nom de machine cliente',type=str, nargs='?', default=socket.gethostname())
	
	args = parser.parse_args()
			
	
	listAccount = []
	listIP = []
	user = ""
	password = ""

	
	if (args.user != None and args.password != None and args.ip != None) or (args.f != None):
		#traitement des entrées en lignes de commande
		if (args.user != None and args.password != None ):
		
			userList = [args.user]
			passwordList = [args.password]
			listAccount = getSMBaccountFromInputForUser("", userList, passwordList, False)
			for i in range(0, len(args.ip)):
				t = getListOfIPsFromInputForUser("", args.ip[i])
				listIP += [t]
			addEntrieAccessList(args.user, listIP)
		# Traitement des entrées du fichier
		if (args.f != None ):
			listAccount += getSMBaccountFromInputForUser(args.f)
			i = 0
			for user,password in listAccount:
				listIP = getListOfIPsFromInputForUser(args.f, "", user)
				listIP = [listIP]
				addEntrieAccessList(user, listIP)
				i += 1
			
	#Si aucun paramètre n'est entré (en ligne de commande ou dans un fichier) une erreur se lève
	elif(args.f == None) :	
		if (args.user == None) : 
			print("Il manque le login.")
		if (args.password == None):
			print("Il manque le mot de passe.")
		if (args.ip == None):
			print("Il manque la liste d'hôte")
		if (args.f == None):
			print("Il n'y a pas de chemin vers un fichier indiqué") 
		exit(1)
	
	
	#Création de la base de mot de passe
	userList = []
	passwordList = []
	for user,password in listAccount:
		userList.append(user)
		passwordList.append(password)
	
	createPasswordBase(userList, passwordList)
	
	#Activation du mode verbose
	if args.verbose :
		global verboseOption
		verboseOption = True
		
	#Activation du scan de contenu
	if args.scancontent :
		global scanContentOption
		scanContentOption = True
	
	
	global regexFileNameList
	
	if args.regexoffice:
		regexFileNameList.append(regex_office)
	
	if args.regextxt:
		regexFileNameList.append(regex_txt)
	
	if args.regexfilename != None:
		for it in args.regexfilename:
			regexFileNameList.append(it)
	
	if regexFileNameList == []:
		print("Il n'y a pas de regex vers des noms de fichiers")
		sys.exit(1)
	
	
	global regexContentList
	if args.regexcontent != None:
		for it in args.regexcontent:
			regexContentList.append(it)
	elif scanContentOption:
		print("Il n'y a pas de regex sur des contenus de fichiers")
		sys.exit(1)
	

	
	global depth
	if args.maxdepth :
		depth = args.maxdepth
	
	global output_filename
	if args.output:
		output_filename = args.output

	global searchlastwriterOption
	if args.searchlastwriter:
		searchlastwriterOption = args.searchlastwriter
		
	global use_ntlm_v2
	if args.ntlmv1:
		use_ntlm_v2 = False
	else:
		use_ntlm_v2 = True
	
	global portSMB
	if args.portsmb:
		portSMB = args.portsmb
		
	global domain_name
	if args.domainname:
		domain_name = args.domainname
		
	global client_machine_name
	if args.clientmachinename:
		client_machine_name = args.clientmachinename
	

#Retourne le hostname correspondant à une adresse IP.
# En cas d'erreur retourne une chaine vide
def nslookup_reverse(ip):
	line = ""
	try:
		reversed_dns = socket.gethostbyaddr(ip)
		line = reversed_dns[0]
		line = line.split('.')[0]
	except socket.herror:
		pass
	return line


# A partir des arguments rentrés en ligne de commande.
# Cette fonction scanne les IPs en entrées et recherche
# des fichiers correspondants à des motifs sur des partages
# réseaux. A noter que les crédentials associés à un range d'ip
# sont aussi donné en paramètre.
# Par défaut les résultats sont écrits sur la sortie standard.
# Sinon dans un fichier de sortie csv.
# Input:
	# argument en ligne de commande
# Output:
	# résultat du scan sur un fichier de sortie (ou la sortie de standard)
def scan():

	global AccessList
	global verboseOption
	global scanContentOption
	global output_filename
	global searchlastwriterOption
	global client_machine_name
	global regexExtensionList
	global regexContentList
	global portSMB
	global use_ntlm_v2
	global domain_name
	global depth
	
	path = Path(output_filename)
	
	for username in AccessList:
		IP_list = getListOfIPFromUserAccess(username)
		
		scan_result = []
		
		
		
		for server_ip in IP_list:
				if(not isPortOpen(server_ip, portSMB)):
					continue
				
				server_name = nslookup_reverse(server_ip)
				
				
				if(not isRequiredAuthentication(server_ip, portSMB)):
					password = "vide"
				else:
					password = givePasswordFromUser(username)
				
				domain_name = 'KNET.INTRA'
				
				if(scanContentOption):
					scan_result = SearchFileContent(username, password, client_machine_name, server_name, server_ip, regexFileNameList,
					verboseOption, regexContentList, portSMB, domain_name, use_ntlm_v2, searchlastwriterOption, depth)
				else:
					scan_result = SearchFileName(username, password, client_machine_name, server_name, server_ip, regexFileNameList, verboseOption,
					portSMB, domain_name, use_ntlm_v2, searchlastwriterOption, depth)
				
				if scan_result == []:
					continue
				else:
					try:
						f = open(path, 'w', newline='')

						if(searchlastwriterOption):
							fieldnames = ['ip', 'hostname', 'create_time', 'owner', 'name', 'path', 'length', 'last_user',
							'regex_fileName']
						else:
							fieldnames = ['ip', 'hostname', 'create_time', 'owner', 'name', 'path', 'length', 'regex_fileName']
						

						if(scanContentOption):
							fieldnames += ['MatchesContent']

							
						writer = csv.DictWriter(f, fieldnames=fieldnames)
						writer.writeheader()
						
						for it in scan_result:
							writer.writerow(it)
						
						if password == "vide":
							f.writelines("Aucune authentification requise")
						
						f.close()
						
					except csv.Error as e:
						print(e)
						sys.exit('file {}, line {}: {}'.format(f, writer.line_num, e))
	
					
	
		
	

#************************
#*****MAIN***************
#************************

handleArgument()
scan()
sys.exit()
